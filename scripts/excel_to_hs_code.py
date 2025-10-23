import openpyxl
import json
import sys
import re
from pathlib import Path

def read_excel_file(file_path, sheet_name=None):
    """Đọc file Excel"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        # Nếu không chỉ định sheet, dùng sheet đầu tiên
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                print(f"⚠️  Sheet '{sheet_name}' không tìm thấy. Các sheet có sẵn: {workbook.sheetnames}")
                sheet_name = workbook.sheetnames[0]
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        
        print(f"📄 Đang đọc sheet: '{worksheet.title}'")
        
        data = []
        # Bắt đầu từ hàng 20 (dữ liệu thực tế bắt đầu từ đây)
        for row in worksheet.iter_rows(min_row=20, values_only=False):
            row_data = {}
            for cell in row:
                row_data[cell.column_letter] = cell.value or ''
            data.append(row_data)
        
        print(f"✓ Đọc thành công {len(data)} hàng từ Excel\n")
        return data
    except Exception as error:
        print(f"❌ Lỗi đọc file Excel: {error}")
        sys.exit(1)

def extract_section_number(description):
    """Lấy số Section từ mô tả"""
    if not description:
        return None
    match = re.search(r'Chương\s+(\d+)', str(description))
    return match.group(1) if match else None

def extract_chapter_number(code):
    """Lấy số Chapter từ mã (2 chữ số đầu)"""
    if not code:
        return None
    code_str = str(code).strip()
    # Lấy 2 chữ số đầu
    if len(code_str) >= 2 and code_str[:2].isdigit():
        return code_str[:2]
    return None

def extract_parent_codes(code):
    """Trích xuất mã parent từ mã HS"""
    if not code:
        return {}
    
    code_str = str(code).strip()
    
    # Kiểm tra mã HS có đủ 4 hoặc 6 chữ số không
    if len(code_str) < 4 or not code_str[:6].replace('00', '').isdigit():
        return {}
    
    # Chuẩn hoá thành 6 chữ số
    if len(code_str) == 4:
        full_code = code_str + '00'
    elif len(code_str) >= 6:
        full_code = code_str[:6]
    else:
        full_code = code_str.ljust(6, '0')
    
    return {
        "chapter": full_code[:2],          # 2 chữ số đầu (XX0000)
        "heading": full_code[:4],          # 4 chữ số đầu (XXYY00)
        "subheading": full_code[:6]        # 6 chữ số (XXYYZZ)
    }

def build_hs_hierarchy(code):
    """Xây dựng cây phân cấp HS Code"""
    if not code:
        return None
    
    code_str = str(code).strip()
    if len(code_str) < 2:
        return None
    
    # Chuẩn hoá thành 6 chữ số
    if len(code_str) == 4:
        full_code = code_str + '00'
    elif len(code_str) >= 6:
        full_code = code_str[:6]
    else:
        full_code = code_str.ljust(6, '0')
    
    return {
        "level_1_chapter": full_code[:2],      # Chương (2 số)
        "level_2_heading": full_code[:4],      # Heading (4 số)
        "level_3_subheading": full_code[:6],   # Subheading (6 số)
        "full_code": full_code
    }

def extract_sections_and_chapters(excel_data):
    """Trích xuất Section và Chapter"""
    sections = {}
    chapters = {}
    
    for row in excel_data:
        col_c = row.get('C', '')
        col_h = row.get('H', '')
        col_d = row.get('D', '')
        
        if col_c and 'Chương' in str(col_c):
            match = re.search(r'Chương\s+(\d+)', str(col_c))
            if match:
                section_num = match.group(1)
                section_key = f"Section_{section_num}"
                
                if section_key not in sections:
                    sections[section_key] = {
                        "number": section_num,
                        "viName": str(col_c),
                        "enName": str(col_h),
                        "description": str(col_d)
                    }
        
        col_e = row.get('E', '')
        col_f = row.get('F', '')
        
        if col_e and any(word in str(col_e).lower() for word in ['animals', 'động vật', 'live']):
            chapter_key = str(col_e).strip()
            
            if chapter_key not in chapters:
                chapters[chapter_key] = {
                    "viName": str(col_d),
                    "enName": str(col_e),
                    "description": str(col_f)
                }
    
    return sections, chapters

def transform_to_hs_code(excel_data):
    """Chuyển đổi dữ liệu sang định dạng HS Code với hierarchy
    
    Ánh xạ cột:
    - E: Level
    - F: Mã hàng (HS Code)
    - G: Description Tiếng Việt
    - H: Description Tiếng Anh
    
    Logic nối description parent:
    - Lấy description của tất cả level cha từ level 0 đến level hiện tại
    - Nối theo thứ tự: Level0 | Level1 | Level2 | Level3 | ...
    """
    hs_codes = []
    level_cache = {}  # Cache description theo level
    
    for row in excel_data:
        col_f = row.get('F', '')
        col_e = row.get('E', '')
        
        # Lấy thông tin từ các cột
        code = str(col_f).strip() if col_f else ""
        level = str(col_e).strip() if col_e else ""
        vi_desc = str(row.get('G', '')).strip()        # Description Tiếng Việt
        en_desc = str(row.get('H', '')).strip()        # Description Tiếng Anh
        
        # Convert level to int để dễ so sánh
        try:
            level_int = int(level) if level else -1
        except:
            level_int = -1
        
        # Lọc hàng có mã HS và bỏ các hàng header/trống
        if code and code[0].isdigit():
            hierarchy = build_hs_hierarchy(code)
            parent_codes = extract_parent_codes(code)
            
            # Xóa cache các level > level hiện tại
            if level_int >= 0:
                keys_to_delete = [k for k in level_cache.keys() if k > level_int]
                for k in keys_to_delete:
                    del level_cache[k]
            
            # Cập nhật cache description của level hiện tại
            level_cache[level_int] = {
                'vi': vi_desc,
                'en': en_desc
            }
            
            # Nối description từ tất cả level cha
            vi_parts = []
            en_parts = []
            for lv in sorted(level_cache.keys()):
                if level_cache[lv]['vi']:
                    vi_parts.append(level_cache[lv]['vi'])
                if level_cache[lv]['en']:
                    en_parts.append(level_cache[lv]['en'])
            
            full_vi_desc = " | ".join(vi_parts)
            full_en_desc = " | ".join(en_parts)
            
            hs_code_entry = {
                "code": code,
                "level": level,
                "viDescription": vi_desc,
                "enDescription": en_desc,
                "viDescriptionFull": full_vi_desc,
                "enDescriptionFull": full_en_desc,
                "section": extract_section_number(vi_desc),
                "chapter": extract_chapter_number(code),
                "hierarchy": hierarchy,
                "parentCodes": parent_codes
            }
            
            hs_codes.append(hs_code_entry)
        else:
            # Hàng không có code - chỉ cập nhật cache
            if level_int >= 0 and vi_desc and en_desc:
                # Xóa cache các level > level hiện tại
                keys_to_delete = [k for k in level_cache.keys() if k > level_int]
                for k in keys_to_delete:
                    del level_cache[k]
                
                # Cập nhật cache
                level_cache[level_int] = {
                    'vi': vi_desc,
                    'en': en_desc
                }
    
    return hs_codes

def save_to_json(data, output_path):
    """Lưu dữ liệu sang JSON"""
    try:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"✓ Lưu thành công: {output_path}")
    except Exception as error:
        print(f"❌ Lỗi lưu file JSON: {error}")
        sys.exit(1)

def main():
    """Hàm chính"""
    input_file = sys.argv[1] if len(sys.argv) > 1 else './data/hs-code.xlsx'
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    output_dir = sys.argv[3] if len(sys.argv) > 3 else './public/data'
    
    print("🔄 Bắt đầu trích xuất với HS Code hierarchy...\n")
    
    # Đọc Excel
    excel_data = read_excel_file(input_file, sheet_name)
    
    # Trích xuất dữ liệu
    hs_code_data = transform_to_hs_code(excel_data)
    
    # Lưu file
    save_to_json(hs_code_data, Path(output_dir) / 'hs-codes.json')
    
    # Lưu dữ liệu kết hợp
    combined = {
        "metadata": {
            "totalCodes": len(hs_code_data),
            "source": input_file,
            "sheet": sheet_name
        },
        "codes": hs_code_data
    }
    save_to_json(combined, Path(output_dir) / 'hs-code-complete.json')
    
    print(f"\n✅ Trích xuất hoàn tất:")
    print(f"   • HS Codes: {len(hs_code_data)} bản ghi")
    print(f"   • Output: {Path(output_dir) / 'hs-codes.json'}")

if __name__ == "__main__":
    main()
